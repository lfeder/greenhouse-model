"""Generate Excel snapshot of the expansion model for offline review.

Runs 4 scenarios (current settings, No-Expansion baseline, 3-NO PH preset, 4-FULL PH preset)
and writes everything from /api/compute into separate sheets.

Usage:  python export_excel.py
Output: expansion_export_<timestamp>.xlsx
"""
import json
import sys
from datetime import datetime
from pathlib import Path

import main
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DIR = Path(__file__).parent

# --- Build presets (mirror frontend BUILD_PRESETS) ---
PRESETS = {
    "3-NO": dict(newKAcres=1.4, newJAcres=1.4, newEAcres=1.4, newLAcres=0, newTAcres=0,
                 packhouseAcres=0,   packhouseCostPerAc=4_000_000, housingPods=2, debug=False),
    "4-FULL": dict(newKAcres=1.4, newJAcres=1.4, newEAcres=2.8, newLAcres=0, newTAcres=0,
                   packhouseAcres=0.5, packhouseCostPerAc=4_000_000, housingPods=2, debug=False),
}

HDR_FONT = Font(bold=True, color="000000")
HDR_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
SECTION_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
SECTION_FONT = Font(bold=True)


def load_settings():
    return json.loads((DIR / "settings.json").read_text(encoding="utf-8"))


def write_header(ws, row, labels):
    for col, lbl in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=lbl)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
    return row + 1


def write_section(ws, row, title, span):
    c = ws.cell(row=row, column=1, value=title)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    for col in range(2, span + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    return row + 1


def write_row(ws, row, label, values, fmt=None):
    ws.cell(row=row, column=1, value=label)
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=i + 2, value=v)
        if fmt and isinstance(v, (int, float)):
            c.number_format = fmt
    return row + 1


def autosize(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 12
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                ln = len(str(v))
                if ln > max_len:
                    max_len = ln
        ws.column_dimensions[letter].width = min(max_len + 2, 28)


def write_scenario(wb, name, d, settings):
    """Write a full scenario to one or more sheets."""
    Y = d["years"]
    years_hdr = [f"'{str(y)[-2:]}" for y in Y]

    # ── 1. KPIs ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet(f"{name} KPIs")
    write_header(ws, 1, ["Metric", "Value", "Detail"])
    last_idx = len(Y) - 1 - 2  # _DN=2 in frontend
    steady_rev = d["rev"][last_idx]
    steady_op = d["op_inc"][last_idx]
    capex = d["total_capex"] + d["shared_capex"]
    startup = d.get("startup_capital", 0)
    bank = d.get("bank_total", 0)
    jjb = d.get("jjb_equity", 0)
    bridge = (d.get("partners_no_exp") or {}).get("guarantee_total", 0) or 0
    combined_ds = [d["loan_total_ds"][i] + d["expansion_ds"][i] for i in range(len(Y))]
    peak = max(combined_ds)
    rows = [
        ("$ Uses (capex + startup)", capex + startup),
        ("  Crop capex", d["total_capex"]),
        ("  Shared infra capex", d["shared_capex"]),
        ("  Startup capital", startup),
        ("$ Sources (bank + JJB)", bank + jjb),
        ("  Bank loan", bank),
        ("  JJB equity", jjb),
        ("  Bridge loan", bridge),
        (f"Steady Rev ({Y[last_idx]})", steady_rev),
        (f"Steady Op Inc ({Y[last_idx]})", steady_op),
        ("Steady Op Margin", steady_op / steady_rev if steady_rev else 0),
        ("Peak debt service", peak),
        ("Total IRR (%)", d["total_irr"]),
        ("Total Unlevered (%)", d.get("total_unlev", 0)),
    ]
    r = 2
    for lbl, v in rows:
        ws.cell(row=r, column=1, value=lbl)
        c = ws.cell(row=r, column=2, value=v)
        if isinstance(v, float):
            c.number_format = "#,##0.00" if "%" in lbl or "Margin" in lbl else "$#,##0"
        else:
            c.number_format = "$#,##0"
        r += 1
    autosize(ws)

    # ── 2. Revenue & Expense ──────────────────────────────────────────────────
    ws = wb.create_sheet(f"{name} Rev-Exp")
    r = write_header(ws, 1, ["Row"] + years_hdr)
    r = write_section(ws, r, "Revenue", len(Y) + 1)
    if d.get("rev_rows"):
        for label, data in d["rev_rows"]:
            r = write_row(ws, r, label, data, "$#,##0")
    r = write_row(ws, r, "Total Revenue", d["rev"], "$#,##0")
    r = write_section(ws, r, "Expenses", len(Y) + 1)
    if d.get("exp_rows"):
        for label, data in d["exp_rows"]:
            r = write_row(ws, r, label, data, "$#,##0")
    r = write_row(ws, r, "Total Expenses", d["exp"], "$#,##0")
    r = write_row(ws, r, "Operating Income", d["op_inc"], "$#,##0")
    autosize(ws)

    # ── 3. Ownership ──────────────────────────────────────────────────────────
    ws = wb.create_sheet(f"{name} Ownership")
    r = write_header(ws, 1, ["Partner"] + years_hdr)
    for p in ["EB", "JS", "JJB"]:
        vals = [o.get(p, 0) for o in d["ownership"]]
        r = write_row(ws, r, p + " %", vals, "0.00")
    if d.get("ownership_detail"):
        r += 1
        r = write_section(ws, r, "Ownership Detail", len(Y) + 1)
        for row in d["ownership_detail"]:
            label = row[0]
            vals = row[1:]
            fmt = "$#,##0" if "($)" in label else "0.00"
            r = write_row(ws, r, label, vals, fmt)
    autosize(ws)

    # ── 4. Waterfall (with expansion) ─────────────────────────────────────────
    ws = wb.create_sheet(f"{name} Waterfall")
    r = write_header(ws, 1, ["Row"] + years_hdr)
    r = write_row(ws, r, "Operating Income", d["op_inc"], "$#,##0")
    if d.get("jjb_startup_line"):
        r = write_row(ws, r, "JJB Startup Capital (add back)", d["jjb_startup_line"], "$#,##0")
    r = write_row(ws, r, "Existing DS", [-v for v in d["loan_total_ds"]], "$#,##0")
    r = write_row(ws, r, "Expansion DS", [-v for v in d["expansion_ds"]], "$#,##0")
    r = write_row(ws, r, "Capex Reserve", [-v for v in d["capex_res"]], "$#,##0")
    r = write_row(ws, r, "Distributable Cash", d["distrib_cash"], "$#,##0")
    r = write_section(ws, r, "Waterfall", len(Y) + 1)
    r = write_row(ws, r, "Tax", [-v for v in d["tax_cash"]], "$#,##0")
    r = write_row(ws, r, "Tier 0 (EB)", [-v for v in d["tier0_actual"]], "$#,##0")
    r = write_row(ws, r, "Tier 1 (Partners)", [-v for v in d["tier1_actual"]], "$#,##0")
    r = write_row(ws, r, "PE/DD Repayment", [-v for v in d["pedd_payments"]], "$#,##0")
    r = write_row(ws, r, "Tier 2 (Pro-Rata)", [-v for v in d["tier2"]], "$#,##0")
    if any(v > 50 for v in d.get("tier1_shortfall", [])):
        r = write_row(ws, r, "T1 Shortfall → DD", [-v for v in d["tier1_shortfall"]], "$#,##0")
    # Also include no-expansion waterfall for comparison if available
    ne = (d.get("partners_no_exp") or {}).get("_waterfall")
    if ne:
        r += 2
        r = write_section(ws, r, "── No-Expansion Waterfall (baseline) ──", len(Y) + 1)
        r = write_row(ws, r, "Operating Income (no-exp)", ne["op_inc"], "$#,##0")
        r = write_row(ws, r, "Existing DS (no-exp)", [-v for v in ne["loan_total_ds"]], "$#,##0")
        r = write_row(ws, r, "Capex Reserve (no-exp)", [-v for v in ne["capex_res"]], "$#,##0")
        r = write_row(ws, r, "Distributable Cash (no-exp)", ne["distrib_cash"], "$#,##0")
        r = write_row(ws, r, "Tax (no-exp)", [-v for v in ne["tax_cash"]], "$#,##0")
        r = write_row(ws, r, "Tier 0 (no-exp)", [-v for v in ne["tier0_actual"]], "$#,##0")
        r = write_row(ws, r, "Tier 1 (no-exp)", [-v for v in ne["tier1_actual"]], "$#,##0")
        r = write_row(ws, r, "PE/DD (no-exp)", [-v for v in ne["pedd_payments"]], "$#,##0")
        r = write_row(ws, r, "Tier 2 (no-exp)", [-v for v in ne["tier2"]], "$#,##0")
    autosize(ws)

    # ── 5. Partner Cash ───────────────────────────────────────────────────────
    ws = wb.create_sheet(f"{name} Partner Cash")
    r = write_header(ws, 1, ["Row"] + years_hdr)
    for p in ["EB", "JS", "JJB"]:
        pd = d.get("partners", {}).get(p)
        if not pd:
            continue
        r = write_section(ws, r, p, len(Y) + 1)
        r = write_row(ws, r, "Tax Dist", pd["tax_dist"], "$#,##0")
        if pd.get("tier_0"):
            r = write_row(ws, r, "Tier 0", pd["tier_0"], "$#,##0")
        r = write_row(ws, r, "Tier 1", pd["tier_1"], "$#,##0")
        if pd.get("pedd"):
            r = write_row(ws, r, "PE/DD", pd["pedd"], "$#,##0")
        r = write_row(ws, r, "Tier 2", pd["tier_2"], "$#,##0")
        if pd.get("equity_buy_sell"):
            r = write_row(ws, r, "Equity Buy/Sell", pd["equity_buy_sell"], "$#,##0")
        r = write_row(ws, r, f"Total {p}", pd["total"], "$#,##0")
        # After-tax
        ex_tax = [pd["total"][i] - pd["tax_dist"][i] for i in range(len(Y))]
        r = write_row(ws, r, f"After-Tax Cash ({p})", ex_tax, "$#,##0")
        # No-exp comparison
        ne_arr = (d.get("partners_no_exp") or {}).get(p)
        if ne_arr:
            r = write_row(ws, r, "No-Expansion", ne_arr, "$#,##0")
            diff = [pd["total"][i] - ne_arr[i] for i in range(len(Y))]
            r = write_row(ws, r, "Expansion Impact", diff, "$#,##0")
    autosize(ws)

    # ── 6. Tax Detail ─────────────────────────────────────────────────────────
    ws = wb.create_sheet(f"{name} Tax")
    r = write_header(ws, 1, ["Row"] + years_hdr)
    r = write_row(ws, r, "Operating Income", d["op_inc"], "$#,##0")
    r = write_row(ws, r, "Total Interest", [-v for v in d.get("total_interest", d["loan_total_int"])], "$#,##0")
    r = write_row(ws, r, "Taxable Income", d["taxable_inc"], "$#,##0")
    r = write_section(ws, r, "Federal", len(Y) + 1)
    r = write_row(ws, r, "Bonus Depreciation", [-v for v in d["dep"]["fed"]], "$#,##0")
    fed_taxable = [td["fed_taxable"] for td in d["tax_detail"]]
    r = write_row(ws, r, "Fed Taxable", fed_taxable, "$#,##0")
    r = write_row(ws, r, "EB Share", [td["eb_fed"] for td in d["tax_detail"]], "$#,##0")
    r = write_row(ws, r, "NOL Used", [-td["nol_used"] for td in d["tax_detail"]], "$#,##0")
    r = write_row(ws, r, "NOL Generated", [td["nol_generated"] for td in d["tax_detail"]], "$#,##0")
    r = write_row(ws, r, "NOL Balance", [td["nol_remaining"] for td in d["tax_detail"]], "$#,##0")
    r = write_row(ws, r, "EB Net Taxable", [td["eb_net_fed"] for td in d["tax_detail"]], "$#,##0")
    r = write_row(ws, r, "Fed Tax (EB)", [-td["fed_tax"] for td in d["tax_detail"]], "$#,##0")
    r = write_row(ws, r, "Fed Tax (entity)", [-td["fed_dist"] for td in d["tax_detail"]], "$#,##0")
    r = write_section(ws, r, "Hawaii", len(Y) + 1)
    r = write_row(ws, r, "State Depreciation", [-v for v in d["dep"]["state"]], "$#,##0")
    r = write_row(ws, r, "HI Taxable", [td["hi_taxable"] for td in d["tax_detail"]], "$#,##0")
    r = write_row(ws, r, "EB HI Taxable", [td["eb_hi"] for td in d["tax_detail"]], "$#,##0")
    r = write_row(ws, r, "HI Tax (EB)", [-td["hi_tax"] for td in d["tax_detail"]], "$#,##0")
    r = write_row(ws, r, "HI Tax (entity)", [-td["hi_dist"] for td in d["tax_detail"]], "$#,##0")
    r = write_row(ws, r, "Tax Liability", d["tax_liab"], "$#,##0")
    r = write_row(ws, r, "Tax Cash Dist", [-v for v in d["tax_cash"]], "$#,##0")
    autosize(ws)

    # ── 7. Crop IRRs ──────────────────────────────────────────────────────────
    if d.get("crop_irrs"):
        ws = wb.create_sheet(f"{name} Crop IRRs")
        r = write_header(ws, 1, ["Key", "Label", "Lev IRR (%)", "Unlev (%)", "Equity ($)",
                                  "Base Rev", "Base Exp", "Op Inc", "CapEx"])
        for ci in d["crop_irrs"]:
            ud = ci.get("unlev_detail", {})
            ws.cell(row=r, column=1, value=ci.get("key"))
            ws.cell(row=r, column=2, value=ci.get("label"))
            ws.cell(row=r, column=3, value=ci.get("irr"))
            ws.cell(row=r, column=4, value=ci.get("unlev"))
            ws.cell(row=r, column=5, value=ci.get("equity"))
            ws.cell(row=r, column=6, value=ud.get("base_rev"))
            ws.cell(row=r, column=7, value=ud.get("base_exp"))
            ws.cell(row=r, column=8, value=ud.get("op_inc"))
            ws.cell(row=r, column=9, value=ud.get("capex"))
            for col in [5, 6, 7, 8, 9]:
                ws.cell(row=r, column=col).number_format = "$#,##0"
            for col in [3, 4]:
                ws.cell(row=r, column=col).number_format = "0.0"
            r += 1
        autosize(ws)


def write_scenario_comparison(wb, scenarios, constants):
    """Cross-scenario comparison sheets."""
    # Get years from any scenario
    any_d = next(iter(scenarios.values()))
    Y = any_d["years"]
    years_hdr = [f"'{str(y)[-2:]}" for y in Y]

    # ── Scenario Comparison ($M) ─────────────────────────────────────────────
    ws = wb.create_sheet("Scenario Compare")
    r = write_header(ws, 1, ["Scenario / Row"] + years_hdr)
    resol = constants.get("resolution_data", {})

    # Resolution
    r = write_section(ws, r, "RESOLUTION (hardcoded baseline)", len(Y) + 1)
    resol_yrs_offset = 0  # PDF data 2026-2032, model years 2026-2035
    if resol:
        r = write_row(ws, r, "EBITDA (with-tax basis)", resol["EBITDA"] + [None, None, None], "$#,##0")
        for p in ["EB", "JS", "JJB"]:
            r = write_row(ws, r, f"{p} (with tax)", resol[p]["withTax"] + [None, None, None], "$#,##0")
            r = write_row(ws, r, f"{p} (after tax)", resol[p]["noTax"] + [None, None, None], "$#,##0")

    # Computed scenarios
    for name, d in scenarios.items():
        r = write_section(ws, r, name.upper(), len(Y) + 1)
        if name == "No Expansion":
            ne = (d.get("partners_no_exp") or {}).get("_waterfall")
            if ne:
                r = write_row(ws, r, "Revenue", ne.get("rev", d["rev"]), "$#,##0")
                r = write_row(ws, r, "EBITDA", ne["op_inc"], "$#,##0")
            tax = (d.get("partners_no_exp") or {}).get("_tax", {})
            for p in ["EB", "JS", "JJB"]:
                arr = (d.get("partners_no_exp") or {}).get(p, [0] * len(Y))
                r = write_row(ws, r, f"{p} (with tax)", arr, "$#,##0")
                no_tax = [arr[i] - (tax.get(p, [0] * len(Y))[i]) for i in range(len(Y))]
                r = write_row(ws, r, f"{p} (after tax)", no_tax, "$#,##0")
        else:
            r = write_row(ws, r, "Revenue", d["rev"], "$#,##0")
            r = write_row(ws, r, "EBITDA", d["op_inc"], "$#,##0")
            for p in ["EB", "JS", "JJB"]:
                pd = d.get("partners", {}).get(p, {})
                total = pd.get("total", [0] * len(Y))
                tax = pd.get("tax_dist", [0] * len(Y))
                r = write_row(ws, r, f"{p} (with tax)", total, "$#,##0")
                no_tax = [total[i] - tax[i] for i in range(len(Y))]
                r = write_row(ws, r, f"{p} (after tax)", no_tax, "$#,##0")
    autosize(ws)

    # ── Cash Flow & Stake Value Summary ──────────────────────────────────────
    ws = wb.create_sheet("CF + Stake Summary")
    r = write_header(ws, 1, ["Scenario", "EB", "JS", "JJB", "Detail"])
    r = write_section(ws, r, "Cum After-Tax CF ('26–'29)", 5)

    # Resolution
    if resol:
        resol_cum = {p: sum(resol[p]["noTax"][:4]) for p in ["EB", "JS", "JJB"]}
        r = write_row(ws, r, "Resolution", [resol_cum["EB"], resol_cum["JS"], resol_cum["JJB"], "Hardcoded PDF"], "$#,##0")

    # Computed
    idx2033 = Y.index(2033) if 2033 in Y else len(Y) - 3
    for name, d in scenarios.items():
        if name == "No Expansion":
            partners_arr = d.get("partners_no_exp", {})
            tax = partners_arr.get("_tax", {})
            cum = {}
            for p in ["EB", "JS", "JJB"]:
                arr = partners_arr.get(p, [0] * len(Y))
                cum[p] = sum(arr[i] - tax.get(p, [0] * len(Y))[i] for i in range(4))
        else:
            cum = {}
            for p in ["EB", "JS", "JJB"]:
                pd = d.get("partners", {}).get(p, {})
                total = pd.get("total", [0] * len(Y))
                tax_d = pd.get("tax_dist", [0] * len(Y))
                cum[p] = sum(total[i] - tax_d[i] for i in range(4))
        r = write_row(ws, r, name, [cum["EB"], cum["JS"], cum["JJB"], ""], "$#,##0")

    r = write_section(ws, r, "Stake Value '33 (EBITDA × multiple × own%)", 5)
    # Multiples per scenario
    multiples = {"Resolution": 7, "No Expansion": 7, "3 blk · No PH": 8.5, "4 blk · Full PH": 10}
    # Resolution
    if resol:
        ebitda33 = resol["EBITDA"][-1] * 1.04
        own = {"EB": 27.5, "JS": 22.5, "JJB": 50.0}
        mult = 7
        bv = ebitda33 * mult
        r = write_row(ws, r, f"Resolution", [bv * own["EB"] / 100, bv * own["JS"] / 100, bv * own["JJB"] / 100,
                                              f"EBITDA ${ebitda33/1e6:.2f}M × {mult}x = ${bv/1e6:.2f}M biz"], "$#,##0")
    for name, d in scenarios.items():
        mult = multiples.get(name, 7)
        if name == "No Expansion":
            ne = (d.get("partners_no_exp") or {}).get("_waterfall", {})
            ebitda33 = ne.get("op_inc", [0] * len(Y))[idx2033]
            own = {"EB": 27.5, "JS": 22.5, "JJB": 50.0}
        else:
            ebitda33 = d.get("op_inc", [0] * len(Y))[idx2033]
            own_row = d.get("ownership", [{}])[idx2033] if d.get("ownership") else {}
            own = {p: own_row.get(p, 0) for p in ["EB", "JS", "JJB"]}
        bv = ebitda33 * mult
        r = write_row(ws, r, name,
                       [bv * own["EB"] / 100, bv * own["JS"] / 100, bv * own["JJB"] / 100,
                        f"EBITDA ${ebitda33/1e6:.2f}M × {mult}x = ${bv/1e6:.2f}M biz"], "$#,##0")
    autosize(ws)


def write_settings(wb, settings):
    """Inputs sheet."""
    ws = wb.create_sheet("Inputs", 0)  # Insert at front
    r = write_header(ws, 1, ["Key", "Value"])
    for k in sorted(settings.keys()):
        ws.cell(row=r, column=1, value=k)
        v = settings[k]
        if isinstance(v, (int, float, str, bool)):
            ws.cell(row=r, column=2, value=v)
        else:
            ws.cell(row=r, column=2, value=str(v))
        r += 1
    autosize(ws)


def main_export():
    settings = load_settings()
    constants = json.loads((DIR / "constants.json").read_text(encoding="utf-8"))

    print("Running scenarios...")
    scenarios = {}
    # Current settings
    print("  Current...")
    scenarios["Current"] = main.run_everything(dict(settings))
    # No expansion baseline (already inside Current as partners_no_exp, but easier to run fresh)
    print("  No Expansion...")
    s_ne = dict(settings)
    s_ne.update({"newKAcres": 0, "newJAcres": 0, "newEAcres": 0, "newLAcres": 0, "newTAcres": 0,
                 "packhouseAcres": 0, "housingPods": 0})
    scenarios["No Expansion"] = main.run_everything(s_ne)
    # 3-NO PH
    print("  3 blk · No PH...")
    s3 = dict(settings); s3.update(PRESETS["3-NO"])
    scenarios["3 blk · No PH"] = main.run_everything(s3)
    # 4-FULL PH
    print("  4 blk · Full PH...")
    s4 = dict(settings); s4.update(PRESETS["4-FULL"])
    scenarios["4 blk · Full PH"] = main.run_everything(s4)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Write inputs first
    write_settings(wb, settings)

    # Each scenario gets its own block of sheets
    for name, d in scenarios.items():
        # Sheet name limit is 31 chars; shorten
        short = {"Current": "Cur", "No Expansion": "NoExp", "3 blk · No PH": "3-NO", "4 blk · Full PH": "4-FULL"}[name]
        write_scenario(wb, short, d, settings)

    # Cross-scenario comparison
    write_scenario_comparison(wb, scenarios, constants)

    # Save
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    out = DIR / f"expansion_export_{stamp}.xlsx"
    wb.save(out)
    print(f"\nWrote: {out}")
    print(f"  Size: {out.stat().st_size / 1024:.1f} KB")
    return out


if __name__ == "__main__":
    main_export()
